# main.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import json
import os
import datetime
import zipfile
import textwrap
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import mm
import sys
from openpyxl import load_workbook

def resource_path(relative_path):
    if getattr(sys, 'frozen', False):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

TEST_CASES_FILE = resource_path("test_cases.json")
USERS_FILE = resource_path("users.json")
ATTACHMENTS_DIR = "attachments"
os.makedirs(ATTACHMENTS_DIR, exist_ok=True)

def load_json(file):
    if os.path.exists(file):
        with open(file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_json(file, data):
    with open(file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

class LoginWindow:
    def __init__(self, master, on_success):
        self.master = master
        self.master.title("Login_Test Manager System")
        self.on_success = on_success

        tk.Label(master, text="Name:").grid(row=0, column=0)
        self.name_entry = tk.Entry(master)
        self.name_entry.grid(row=0, column=1)

        tk.Label(master, text="Employee ID:").grid(row=1, column=0)
        self.id_entry = tk.Entry(master, show='*')
        self.id_entry.grid(row=1, column=1)

        tk.Button(master, text="Login", command=self.login).grid(row=2, column=0, columnspan=2)

    def login(self):
        users = load_json(USERS_FILE)
        name = self.name_entry.get().strip()
        emp_id = self.id_entry.get().strip()

        for user in users:
            if user['name'] == name and user['emp_id'] == emp_id:
                self.master.destroy()
                self.on_success(user)
                return

        messagebox.showerror("Login Failed", "Invalid credentials.")

class TestCaseManager:
    def __init__(self, user):
        self.user = user
        self.test_cases = load_json(TEST_CASES_FILE)

        self.root = tk.Tk()
        icon_path = resource_path("app_icon.ico")
        if os.path.exists(icon_path):
            self.root.iconbitmap(icon_path)
        self.root.title("Test Case Manager")
        self.setup_ui()
        self.refresh_table()
        self.root.mainloop()

    def setup_ui(self):
        top_frame = tk.Frame(self.root)
        top_frame.pack(fill='x')

        tk.Label(top_frame, text=f"Logged in as: {self.user['name']} ({self.user['role']})").pack(side='left')
        tk.Button(top_frame, text="Generate Report", command=self.generate_report).pack(side='right')
        if self.user['role'] == 'admin':
            tk.Button(top_frame, text="Import from Excel", command=self.import_from_excel).pack(side='right')
            tk.Button(top_frame, text="Add Test Case", command=self.add_test_case).pack(side='right')

        columns = ("id", "desc", "expected", "status", "comment", "executed_by", "exec_date", "attachment")
        self.tree = ttk.Treeview(self.root, columns=columns, show='headings')
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor='w', stretch=True, width=150)

        vsb = ttk.Scrollbar(self.root, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(self.root, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        self.tree.pack(fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        hsb.pack(side='bottom', fill='x')

        self.tree.bind("<Double-1>", self.update_status_and_comment)

    def refresh_table(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for tc in self.test_cases:
            self.tree.insert('', 'end', values=(
                tc.get('Test Case ID', ''),
                tc.get('Description', ''),
                tc.get('Expected Result', ''),
                tc.get('status', 'Pending'),
                tc.get('comment', ''),
                tc.get('executed_by', ''),
                tc.get('exec_date', ''),
                tc.get('attachment', '')
            ))

    def import_from_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not file_path:
            return

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            header_row = None
            for row in ws.iter_rows(values_only=True):
                if row and any(cell is not None for cell in row):
                    header_row = [str(cell).strip() if cell else "" for cell in row]
                    break

            expected_fields = ["Test Case ID", "Description", "Expected Result"]
            if not all(field in header_row for field in expected_fields):
                messagebox.showerror("Error", "Excel is missing required columns.")
                return

            col_indices = {name: header_row.index(name) for name in expected_fields}
            test_cases = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                case = {}
                for key in expected_fields:
                    cell = row[col_indices[key]] if col_indices[key] < len(row) else ""
                    case[key] = str(cell).strip() if cell is not None else ""
                case.update({"status": "Pending", "comment": "", "executed_by": "", "exec_date": "", "attachment": ""})
                test_cases.append(case)

            self.test_cases = test_cases
            save_json(TEST_CASES_FILE, self.test_cases)
            self.refresh_table()
            messagebox.showinfo("Success", f"Imported {len(test_cases)} test cases from Excel.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to import from Excel:\n{str(e)}")

    # ... rest of the code unchanged ...
    def add_test_case(self):
        win = tk.Toplevel(self.root)
        win.title("Add Test Case")

        fields = ['id', 'description', 'expected']
        entries = {}
        for i, field in enumerate(fields):
            tk.Label(win, text=field).grid(row=i, column=0)
            ent = tk.Entry(win)
            ent.grid(row=i, column=1)
            entries[field] = ent

        def save():
            tc = {
                "Test Case ID": entries['id'].get(),
                "Description": entries['description'].get(),
                "Expected Result": entries['expected'].get(),
                "status": "Pending",
                "comment": "",
                "executed_by": "",
                "exec_date": "",
                "attachment": ""
            }
            self.test_cases.append(tc)
            save_json(TEST_CASES_FILE, self.test_cases)
            win.destroy()
            self.refresh_table()

        tk.Button(win, text="Save", command=save).grid(row=len(fields), columnspan=2)

    def update_status_and_comment(self, event):
        selected = self.tree.selection()
        if not selected:
            return
        item = self.tree.item(selected[0])
        test_id = item['values'][0]

        for tc in self.test_cases:
            if str(tc.get('Test Case ID', '')) == str(test_id):
                win = tk.Toplevel(self.root)
                win.title("Test Case Details")

                def readonly_entry(parent, value):
                    entry = tk.Entry(parent)
                    entry.insert(0, value)
                    entry.config(state='readonly')
                    return entry

                def readonly_text(parent, value):
                    txt = tk.Text(parent, height=4, wrap='word')
                    txt.insert('1.0', value)
                    txt.config(state='disabled')
                    return txt

                row = 0
                tk.Label(win, text="ID:").grid(row=row, column=0, sticky='w')
                readonly_entry(win, tc.get('Test Case ID', '')).grid(row=row, column=1, sticky='ew')

                row += 1
                tk.Label(win, text="Description:").grid(row=row, column=0, sticky='nw')
                readonly_text(win, tc.get('Description', '')).grid(row=row, column=1, sticky='ew')

                row += 1
                tk.Label(win, text="Expected Result:").grid(row=row, column=0, sticky='nw')
                readonly_text(win, tc.get('Expected Result', '')).grid(row=row, column=1, sticky='ew')

                row += 1
                tk.Label(win, text="Status:").grid(row=row, column=0, sticky='w')
                status_cb = ttk.Combobox(win, values=["Pending", "Done", "Deviation"], state="readonly")
                status_cb.set(tc.get('status', 'Pending'))
                status_cb.grid(row=row, column=1, sticky='ew')

                row += 1
                tk.Label(win, text="Comment:").grid(row=row, column=0, sticky='w')
                comment_entry = tk.Entry(win)
                comment_entry.insert(0, tc.get('comment', ''))
                comment_entry.grid(row=row, column=1, sticky='ew')

                row += 1
                tk.Label(win, text="Executed By:").grid(row=row, column=0, sticky='w')
                readonly_entry(win, self.user['name']).grid(row=row, column=1, sticky='ew')

                row += 1
                tk.Label(win, text="Exec Date:").grid(row=row, column=0, sticky='w')
                exec_label = tk.Label(win, text=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                exec_label.grid(row=row, column=1, sticky='w')

                row += 1
                tk.Label(win, text="Attachment:").grid(row=row, column=0, sticky='w')
                attach_label = tk.Label(win, text=tc.get('attachment', ''))
                attach_label.grid(row=row, column=1, sticky='w')

                def browse_file():
                    file_path = filedialog.askopenfilename(filetypes=[("All files", "*.*")])
                    if file_path:
                        file_name = os.path.basename(file_path)
                        dest = os.path.join(ATTACHMENTS_DIR, f"{test_id}_{file_name}")
                        with open(file_path, 'rb') as fsrc:
                            with open(dest, 'wb') as fdst:
                                fdst.write(fsrc.read())
                        tc['attachment'] = dest
                        attach_label.config(text=dest)

                tk.Button(win, text="Attach File", command=browse_file).grid(row=row+1, columnspan=2)

                def save():
                    tc['status'] = status_cb.get()
                    tc['comment'] = comment_entry.get()
                    tc['executed_by'] = self.user['name']
                    tc['exec_date'] = exec_label.cget("text")
                    save_json(TEST_CASES_FILE, self.test_cases)
                    win.destroy()
                    self.refresh_table()

                tk.Button(win, text="Save", command=save).grid(row=row+2, column=1, sticky='e')
                tk.Button(win, text="Close", command=win.destroy).grid(row=row+2, column=0)
                win.columnconfigure(1, weight=1)
                break

    def generate_report(self):
        if not self.test_cases:
            messagebox.showwarning("No Data", "No test cases available.")
            return

        project_code = simpledialog.askstring("Report Name", "Enter project code or report name:")
        if not project_code:
            return

        zip_path = filedialog.asksaveasfilename(
            defaultextension=".zip",
            filetypes=[("ZIP Files", "*.zip")],
            initialfile=f"{project_code}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        )
        if not zip_path:
            return

        try:
            temp_pdf_path = "temp_report.pdf"
            attachments_to_include = []

            c = canvas.Canvas(temp_pdf_path, pagesize=A4)
            width, height = A4
            x_margin = 20
            y = height - 40

            c.setFont("Helvetica-Bold", 14)
            c.drawString(x_margin, y, f"Test Report - {project_code}")
            y -= 30

            c.setFont("Helvetica", 10)
            for idx, tc in enumerate(self.test_cases, 1):
                lines = [
                    f"ID: {tc.get('Test Case ID', '')}",
                    f"Description: {tc.get('Description', '').replace(chr(10), ' ')}",
                    f"Expected: {tc.get('Expected Result', '')}",
                    f"Comment: {tc.get('comment', '')}",
                    f"Executed By: {tc.get('executed_by', '')}",
                    f"Exec Date: {tc.get('exec_date', '')}",
                    f"Attachment: {os.path.basename(tc.get('attachment', '')) if tc.get('attachment') else 'None'}"
                ]

                for line in lines:
                    for subline in textwrap.wrap(line, width=100):
                        if y < 50:
                            c.showPage()
                            y = height - 40
                            c.setFont("Helvetica", 10)
                        c.drawString(x_margin, y, subline)
                        y -= 14

                status = tc.get('status', '').strip().lower()
                status_text = f"Status: {status.capitalize() if status else 'Pending'}"
                if status == 'done':
                    c.setFillColor(colors.green)
                elif status == 'deviation':
                    c.setFillColor(colors.red)
                elif status == 'pending':
                    c.setFillColor(colors.gray)
                else:
                    c.setFillColor(colors.black)

                if y < 50:
                    c.showPage()
                    y = height - 40
                    c.setFont("Helvetica", 10)

                c.drawString(x_margin, y, status_text)
                c.setFillColor(colors.black)
                y -= 20

                c.setStrokeColor(colors.lightgrey)
                c.line(x_margin, y, width - x_margin, y)
                y -= 10

                if tc.get('attachment') and os.path.exists(tc['attachment']):
                    attachments_to_include.append(tc['attachment'])

            c.save()

            with zipfile.ZipFile(zip_path, 'w') as zipf:
                zipf.write(temp_pdf_path, arcname="Test_Report.pdf")
                for file_path in attachments_to_include:
                    arcname = os.path.join("attachments", os.path.basename(file_path))
                    zipf.write(file_path, arcname=arcname)

            os.remove(temp_pdf_path)
            messagebox.showinfo("Success", f"ZIP report saved to:\n{zip_path}")

        except Exception as e:
            import traceback
            messagebox.showerror("Error", f"Failed to generate report:\n{traceback.format_exc()}")

def main():
    def on_login_success(user):
        TestCaseManager(user)

    root = tk.Tk()
    LoginWindow(root, on_login_success)
    root.mainloop()

if __name__ == '__main__':
    main()
